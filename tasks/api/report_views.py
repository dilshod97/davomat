from django.db.models import OuterRef, Subquery
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import Workbook
from django.utils.timezone import make_aware
from django.db.models import Count, Case, When, Value, CharField
import io
from account.models import User
from rest_framework.permissions import IsAuthenticated
from datetime import datetime, time
from ..models import Attendance, Task, Distance
import requests

BOT_TOKEN = "7988185659:AAHkp0AnenS5_P674Tkf47baNJ3uM3azwRU"
BOT_API_URL = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument"


class DailyReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        is_excel = request.query_params.get("is_excel", "false").lower() == "true"
        day = request.query_params.get("day")  # masalan: 2025-10-20
        info_type = request.query_params.get("info_type", "attendance")

        if not day:
            return Response({"error": "Сана киритилмади (day параметри керак)."}, status=400)

        try:
            day_obj = datetime.strptime(day, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Сана формати нотўғри. Масалан: 2025-10-20"}, status=400)

        user = request.user

        if user.as_user == 0:
            all_users = User.objects.filter(id=user.id)
        elif user.as_user == 1:
            if user.sector_id:
                all_users = User.objects.filter(sector_id=user.sector_id)
            else:
                return Response({"error": "Foydalanuvchi sektorga biriktirilmagan."}, status=400)
        elif user.as_user == 2:
            all_users = User.objects.all()

        else:
            return Response({"error": "as_user noto‘g‘ri qiymatga ega."}, status=400)

        data = []

        for user in all_users:
            rec = Attendance.objects.filter(
                user=user, created_at__date=day_obj, info_type=info_type
            ).order_by("-created_at").first()

            if rec:
                izoh = rec.task_description or ""
                holat = ""

                if not rec.timestamp:
                    holat = "Давомат киритмади"
                else:
                    if rec.created_at.time() > time(9, 0):
                        holat = "Кечикди"
                    else:
                        holat = "Ўз вақтида келди"

                    distance = Distance.objects.filter(attendance=rec)

                    other_address = False

                    for i in distance:
                        if i.distance > 1000:
                            other_address = True

                    if other_address:
                        holat = "Шубҳали (жой узоқ)"
                        izoh += " | Киритилган локация ва объект ўртасидаги масофа катта."

                tasks = list(rec.task.values_list("task", flat=True))
                if tasks:
                    task_list = "\n".join([f"{i + 1}. {t}" for i, t in enumerate(tasks)])
                else:
                    task_list = "-"

                data.append({
                    "user": user.get_full_name() if hasattr(user, "get_full_name") else str(user),
                    "position": getattr(user, "position", "-"),
                    "entered_time": rec.timestamp.strftime("%Y-%m-%d %H:%M") if rec.timestamp else "-",
                    "system_time": rec.created_at.strftime("%Y-%m-%d %H:%M"),
                    "distance": f"{rec.distance} м" if hasattr(rec, "distance") and rec.distance else "-",
                    "status": holat,
                    "tasks": task_list,
                    "where_is_it": rec.task_description,
                    "description": izoh if info_type == "attendance" else rec.description,
                })
            else:
                # 🔹 Agar user bu kunga hech narsa kiritmagan bo‘lsa
                data.append({
                    "user": user.get_full_name() if hasattr(user, "get_full_name") else str(user),
                    "position": getattr(user, "position", "-"),
                    "entered_time": "-",
                    "system_time": "-",
                    "distance": "-",
                    "status": "Давомат киритмади",
                    "tasks": '',
                    "where_is_it": "-",
                    "description": "",
                })

        # 🔹 Excel shaklda chiqarish
        if is_excel:
            wb = Workbook()
            ws = wb.active
            ws.title = "Кунлик ҳисобот"

            headers = ["Ходим", "Лавозими", "Келган вақт", "Тизим вақти", "Масофа", "Ҳолат", "Амалга оширадиган ишлари", "Қаердалиги", "Изоҳ"]
            ws.append(headers)

            for row in data:
                ws.append(list(row.values()))

            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)

            if request.user.chat_id:  # foydalanuvchi ID bo'lsa, bot orqali yuboramiz
                files = {'document': ('kunlik_hisobot.xlsx', file_stream,
                                      'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                payload = {'chat_id': request.user.chat_id,
                           'caption': f"📊 {day} учун кунлик ҳисобот"}
                requests.post(BOT_API_URL, data=payload, files=files)

            return Response({"status": True, "msg": "Fayl bot orqali yuborildi."})

        return Response({"kun": day, "hisobot": data})


class PeriodReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        is_excel = request.query_params.get("is_excel", "false").lower() == "true"
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        info_type = request.query_params.get("info_type", "attendance")

        if not start_date or not end_date:
            return Response(
                {"error": "Сана оралиғини киритинг: start_date ва end_date параметрлари керак."},
                status=400,
            )

        try:
            start_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"error": "Сана формати нотўғри. Масалан: start_date=2025-10-01&end_date=2025-10-20"},
                status=400,
            )

        # mavjud holatlar ro‘yxati
        descriptions = [
            "masofaviy",
            "obyekt_hudud",
            "idora",
            "ruxsat_olganman",
            "taatil",
            "tibbiy_korik_obyekt",
            "obyekt_tashkent",
            "kasal",
            "tibbiy_korik_idora",
        ]

        user = request.user

        if user.as_user == 0:
            all_users = User.objects.filter(id=user.id)
        elif user.as_user == 1:
            if user.sector_id:
                all_users = User.objects.filter(sector_id=user.sector_id)
            else:
                return Response({"error": "Foydalanuvchi sektorga biriktirilmagan."}, status=400)
        elif user.as_user == 2:
            all_users = User.objects.all()

        else:
            return Response({"error": "as_user noto‘g‘ri qiymatga ega."}, status=400)
        data = []

        for user in all_users:
            recs = Attendance.objects.filter(
                user=user,
                created_at__date__range=[start_obj, end_obj],
                info_type=info_type,
            ).order_by("created_at")

            if recs.exists():
                jami = recs.count()
                kechikkanlar = recs.filter(created_at__time__gt=time(9, 0)).count()
                shubhali = recs.filter(latitude__isnull=True, longitude__isnull=True).count()

                oxirgisi = recs.last()

                # Amalga oshirilgan ishlar
                tasks = Task.objects.filter(attendance__in=recs).distinct()
                task_list = "\n".join([f"{i + 1}. {t.task}" for i, t in enumerate(tasks)]) if tasks else "-"

                # har bir task_description bo‘yicha sanash
                desc_counts = {}
                for desc in descriptions:
                    desc_counts[desc] = recs.filter(task_description=desc).count()

                row = {
                    "user": user.get_full_name() if hasattr(user, "get_full_name") else str(user),
                    "position": getattr(user, "position", "-"),
                    "jami_kiritilgan": jami,
                    "kechikkanlar": kechikkanlar,
                    "shubhali_holatlar": shubhali,
                    "songgi_kiritilgan_vaqt": oxirgisi.created_at.strftime("%Y-%m-%d %H:%M"),
                    "amalga_oshirilgan_ishlar": task_list,
                }
                row.update(desc_counts)
                data.append(row)
            else:
                # hech narsa kiritmagan foydalanuvchi
                row = {
                    "user": user.get_full_name() if hasattr(user, "get_full_name") else str(user),
                    "position": getattr(user, "position", "-"),
                    "jami_kiritilgan": 0,
                    "kechikkanlar": 0,
                    "shubhali_holatlar": 0,
                    "songgi_kiritilgan_vaqt": "-",
                    "amalga_oshirilgan_ishlar": "-",
                }
                # nol holatlar uchun ham fieldlar bo‘lishi kerak
                for desc in descriptions:
                    row[desc] = 0
                data.append(row)

        # 🔸 Excel eksport
        if is_excel:
            wb = Workbook()
            ws = wb.active
            ws.title = "Давр ҳисоботи"

            headers = [
                "Ходим",
                "Лавозими",
                "Жами киритилган",
                "Кечикканлар сони",
                "Шубҳали ҳолатлар",
                "Сўнгги киритилган вақт",
                "Амалга оширган ишлари",
            ] + descriptions

            ws.append(headers)

            for row in data:
                ws.append([row.get(h, "") for h in row.keys()])

            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)

            if request.user.chat_id:  # foydalanuvchi ID bo'lsa, bot orqali yuboramiz
                files = {'document': ('kunlik_hisobot.xlsx', file_stream,
                                      'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                payload = {'chat_id': request.user.chat_id, 'caption': f"📊 {start_date} - {end_date} учун кунлик ҳисобот"}
                requests.post(BOT_API_URL, data=payload, files=files)

            return Response({"status": True, "msg": "Fayl bot orqali yuborildi."})

        # 🔸 JSON qaytarish
        return Response({
            "davr": f"{start_date} — {end_date}",
            "hisobot": data
        })


class BandlikHisobotAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        day = request.query_params.get("day")
        is_excel = request.query_params.get("is_excel", "false").lower() == "true"
        info_type = request.query_params.get("info_type", "attendance")

        if not day:
            return Response({"error": "day параметри талаб қилинади. Масалан: ?day=2025-10-20"}, status=400)

        try:
            day_obj = datetime.strptime(day, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Сана формати нотўғри. Масалан: ?day=2025-10-20"}, status=400)

        start = make_aware(datetime.combine(day_obj, datetime.min.time()))
        end = make_aware(datetime.combine(day_obj, datetime.max.time()))

        queryset = (
            Attendance.objects
            .filter(timestamp__range=[start, end], info_type=info_type)
            .annotate(
                bandlik_turi=Case(
                    When(task_description="idora", then=Value("Идорада")),
                    When(task_description="obyekt_tashkent", then=Value("Объектда (Тошкент ш.)")),
                    When(task_description="obyekt_hudud", then=Value("Объектда (Ҳудуд)")),
                    When(task_description__in=["tibbiy_korik_obyekt", "tibbiy_korik_idora"], then=Value("Тиббий кўрик")),
                    When(task_description="kasal", then=Value("Касал")),
                    When(task_description="masofaviy", then=Value("Масофавий")),
                    When(task_description="taatil", then=Value("Таътил")),
                    When(task_description="ruxsat_olganman", then=Value("Рухсат олган")),
                    default=Value("Номаълум"),
                    output_field=CharField(),
                )
            )
            .values("bandlik_turi", "task_description")
            .annotate(soni=Count("user", distinct=True))
        )

        # 🔹 Natijalarni dict shaklida tayyorlash
        data = {}
        for item in queryset:
            key = item["task_description"]
            bandlik_turi = item["bandlik_turi"]
            soni = item["soni"]
            data[key] = {"bandlik_turi": bandlik_turi, "soni": soni}

        jami_xodimlar = User.objects.count()

        # 🔹 Tartib (key + label)
        tartib = [
            {"key": "idora", "label": "Идорада"},
            {"key": "obyekt_tashkent", "label": "Объектда (Тошкент ш.)"},
            {"key": "obyekt_hudud", "label": "Объектда (Ҳудуд)"},
            {"key": "tibbiy_korik_obyekt", "label": "Тиббий кўрик"},
            {"key": "kasal", "label": "Касал"},
            {"key": "masofaviy", "label": "Масофавий"},
            {"key": "taatil", "label": "Таътил"},
            {"key": "ruxsat_olganman", "label": "Рухсат олган"},
        ]

        rows = [
            {
                "key": t["key"],
                "bandlik_turi": t["label"],
                "soni": data.get(t["key"], {}).get("soni", 0),
            }
            for t in tartib
        ]

        # 🔹 Jami qator
        rows.append({"key": "jami", "bandlik_turi": "Жами", "soni": jami_xodimlar})

        # 🔹 Excel
        if is_excel:
            wb = Workbook()
            ws = wb.active
            ws.title = "Бандлик ҳисоботи"

            ws.append(["№", "Калит (key)", "Бандлик тури", "Сони"])
            for i, row in enumerate(rows, start=1):
                ws.append([i, row["key"], row["bandlik_turi"], row["soni"]])

            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)

            if request.user.chat_id:  # foydalanuvchi ID bo'lsa, bot orqali yuboramiz
                files = {'document': ('kunlik_hisobot.xlsx', file_stream,
                                      'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                payload = {'chat_id': request.user.chat_id,
                           'caption': f"📊 {day} учун кунлик ҳисобот"}
                requests.post(BOT_API_URL, data=payload, files=files)

            return Response({"status": True, "msg": "Fayl bot orqali yuborildi."})

        # 🔹 JSON shaklida qaytarish
        return Response({
            "day": str(day_obj),
            "hisobot": rows
        })
